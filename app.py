from flask import Flask, render_template, request, jsonify, g
from datetime import datetime
import os, re, socket, sqlite3, subprocess, tempfile

app = Flask(__name__)
app.config['SECRET_KEY']         = 'zebra-label-2024'
app.config['DATABASE']           = os.path.join(os.path.dirname(__file__), 'etiquetas.db')
app.config['UPLOAD_FOLDER']      = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    db = sqlite3.connect(app.config['DATABASE'])
    db.execute('DROP TABLE IF EXISTS productos_catalogo')
    db.execute("""
        CREATE TABLE IF NOT EXISTS configuracion (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS proveedores (
            codigo TEXT PRIMARY KEY,
            nombre TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS lote_productos (
            codigo       TEXT PRIMARY KEY,
            descripcion  TEXT NOT NULL,
            cantidad     INTEGER DEFAULT 1,
            costo_usd    REAL DEFAULT 0,
            precio_venta REAL DEFAULT 0,
            referencia   TEXT DEFAULT ''
        )
    """)
    # Migration: add referencia column to existing databases
    try:
        db.execute("ALTER TABLE lote_productos ADD COLUMN referencia TEXT DEFAULT ''")
    except Exception:
        pass
    db.commit()
    db.close()

# ── MURCIELAGO ────────────────────────────────────────────────────────────────
_MUR_MAP = {'0':'O','1':'M','2':'U','3':'R','4':'C','5':'I','6':'E','7':'L','8':'A','9':'G'}

def a_murcielago(costo_usd):
    import math
    dec    = costo_usd - math.floor(costo_usd)
    entero = math.ceil(costo_usd) if dec > 0.5 else math.floor(costo_usd)
    return ''.join(_MUR_MAP[d] for d in str(max(entero, 0)))

# ── ZPL ───────────────────────────────────────────────────────────────────────
def generar_zpl_item(item, precio_venta, costo_usd, timestamp_unix):
    def trunc(s, n): return (str(s or ''))[:n]

    codigo      = trunc(item.get('codigo',''), 20)
    descripcion = trunc(item.get('descripcion',''), 32)
    cantidad    = int(item.get('cantidad', 1))
    pv_str      = f'{float(precio_venta or 0):,.2f}'
    murcielago  = a_murcielago(float(costo_usd or 0))
    ts          = int(timestamp_unix)
    referencia  = trunc(item.get('referencia',''), 24)
    bc          = re.sub(r'[^A-Z0-9]', '', codigo.upper()) or 'PROD'

    return (
        f'^XA\n'
        f'^PW456\n^LL352\n'
        f'^PQ{cantidad}\n'
        f'^FO4,4^GB448,344,2^FS\n'
        f'^CF0,30\n^FO10,18^FD{codigo}^FS\n'
        f'^CF0,22\n^FO10,56^FD{descripcion}^FS\n'
        f'^FO4,86^GB448,0,1^FS\n'
        f'^CF0,60\n^FO10,100^FDREF {pv_str}^FS\n'
        f'^FO4,172^GB448,0,1^FS\n'
        f'^CF0,20\n^FO10,190^FD{referencia}^FS\n'
        f'^CF0,24\n^FO240,186^FD{murcielago}^FS\n'
        f'^CF0,16\n^FO330,194^FD{ts}^FS\n'
        f'^FO4,208^GB448,0,1^FS\n'
        f'^BY2,3,80\n'
        f'^FO10,214^BCN,80,N,N^FD{bc}^FS\n'
        f'^XZ'
    )

# ── ZPL Estantería ───────────────────────────────────────────────────────────
def _band_zpl(text, y_start, band_h, avail_w=436, x=10, ratio=0.65, max_lines=2):
    """Fit text into a label band at the largest possible font.
    max_lines=1 forces single line (font shrinks to fit); =2 allows word-wrap."""
    text = str(text or '').strip()
    if not text:
        return ''
    n = len(text)
    inner_h = band_h - 14  # 7 px padding top + bottom
    f1 = min(inner_h, int(avail_w / (ratio * n)))
    if f1 >= 50 or max_lines == 1:
        f = max(20, f1)
        y = y_start + (band_h - f) // 2
        return f'^CF0,{f}\n^FO{x},{y}^FD{text}^FS\n'
    # Two lines — split at word boundary closest to midpoint
    mid = n // 2
    spaces = [i for i, c in enumerate(text) if c == ' ']
    cut = min(spaces, key=lambda i: abs(i - mid)) if spaces else mid
    l1, l2 = text[:cut].strip(), text[cut:].strip()
    longer = max(len(l1), len(l2), 1)
    gap = 6
    f2 = max(24, min((inner_h - gap) // 2, int(avail_w / (ratio * longer))))
    cpl = int(avail_w / (ratio * f2))
    l1, l2 = l1[:cpl], l2[:cpl]
    y1 = y_start + (band_h - (2 * f2 + gap)) // 2
    zpl = f'^CF0,{f2}\n^FO{x},{y1}^FD{l1}^FS\n'
    if l2:
        zpl += f'^CF0,{f2}\n^FO{x},{y1 + f2 + gap}^FD{l2}^FS\n'
    return zpl


def generar_zpl_estanteria(codigo, descripcion, referencia, cantidad):
    cant = max(1, int(float(cantidad or 1)))
    return (
        f'^XA\n^PW456\n^LL352\n^PQ{cant}\n'
        f'^FO4,4^GB448,344,2^FS\n'
        f'^FO4,118^GB448,0,2^FS\n'
        f'^FO4,234^GB448,0,2^FS\n'
        + _band_zpl(codigo,      4,   114, max_lines=1)
        + _band_zpl(descripcion, 118, 116, max_lines=2)
        + _band_zpl(referencia,  234, 114, max_lines=1)
        + '^XZ'
    )

# ── Envío ─────────────────────────────────────────────────────────────────────
def enviar_red(zpl, ip, puerto=9100):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(5)
            s.connect((ip, int(puerto)))
            s.sendall(zpl.encode('utf-8'))
        return True, f'Impreso en {ip}:{puerto}'
    except Exception as e:
        return False, str(e)

def enviar_windows(zpl, printer_name):
    try:
        import win32print
        hprinter = win32print.OpenPrinter(printer_name)
        try:
            win32print.StartDocPrinter(hprinter, 1, ("ZPL Label", None, "RAW"))
            try:
                win32print.StartPagePrinter(hprinter)
                win32print.WritePrinter(hprinter, zpl.encode('utf-8'))
                win32print.EndPagePrinter(hprinter)
            finally:
                win32print.EndDocPrinter(hprinter)
        finally:
            win32print.ClosePrinter(hprinter)
        return True, f'Impreso en {printer_name}'
    except Exception as e:
        return False, str(e)

def enviar_usb(zpl, port='COM1'):
    try:
        import serial
        with serial.Serial(port, 9600, timeout=5) as ser:
            ser.write(zpl.encode('utf-8'))
        return True, 'Impreso exitosamente'
    except ImportError:
        return False, 'Instala pyserial: pip install pyserial'
    except Exception as e:
        return False, str(e)

# ── Helpers Excel ─────────────────────────────────────────────────────────────
# Posiciones fijas para "Imprimir Excel" (sin encabezados, datos desde fila 2):
#   A=0 código  B=1 descripción  E=4 cantidad  G=6 costo  K=10 precio
_POS = {'codigo': 0, 'descripcion': 1, 'cantidad': 4, 'costo': 6, 'precio': 10}

def _col_map(header_row):
    return {str(v or '').strip().lower(): i for i, v in enumerate(header_row)}

def _get(row, col_map, *names, default=None):
    for name in names:
        idx = col_map.get(name.lower())
        if idx is not None and idx < len(row):
            return row[idx]
    return default

def _get_pos(row, *names, default=None):
    for name in names:
        idx = _POS.get(name.lower())
        if idx is not None and idx < len(row):
            return row[idx]
    return default

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/windows_printers')
def windows_printers():
    try:
        import win32print
        flags    = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        printers = [p[2] for p in win32print.EnumPrinters(flags, None, 1)]
        return jsonify({'printers': printers})
    except Exception as e:
        return jsonify({'printers': [], 'error': str(e)})

@app.route('/api/catalogo/stats')
def catalogo_stats():
    count = get_db().execute('SELECT COUNT(*) FROM lote_productos').fetchone()[0]
    return jsonify({'count': count})

@app.route('/api/catalogo/buscar')
def catalogo_buscar():
    codigo = request.args.get('codigo','').strip().upper()
    if not codigo:
        return jsonify({'descripcion': None})
    row = get_db().execute(
        'SELECT descripcion, cantidad, costo_usd, precio_venta, referencia FROM lote_productos WHERE codigo=?', (codigo,)
    ).fetchone()
    if row:
        return jsonify({'descripcion': row[0], 'cantidad': row[1],
                        'costo_usd': row[2], 'precio_venta': row[3], 'referencia': row[4] or ''})
    return jsonify({'descripcion': None})

@app.route('/api/productos/importar', methods=['POST'])
def productos_importar():
    if 'excel' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400
    archivo = request.files['excel']
    fname   = archivo.filename.lower()
    if not fname.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx, .xls)'}), 400
    try:
        if fname.endswith('.xlsx'):
            import openpyxl
            wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        else:
            import xlrd
            wb   = xlrd.open_workbook(file_contents=archivo.read())
            ws   = wb.sheet_by_index(0)
            rows = [tuple(ws.row_values(r)) for r in range(ws.nrows)]
    except Exception as e:
        return jsonify({'error': f'Error al leer el Excel: {e}'}), 400

    if not rows:
        return jsonify({'error': 'El archivo está vacío'}), 400
    cm = _col_map(rows[0])
    db = get_db()
    insertados, actualizados = 0, 0
    for row in rows[1:]:
        codigo = str(_get(row, cm, 'codigo') or '').strip().upper()
        desc   = str(_get(row, cm, 'descripcion') or '').strip()
        if not codigo or not desc: continue
        try: costo  = float(_get(row, cm, 'costo', 'costo_usd',    default=0) or 0)
        except: costo = 0.0
        try: precio = float(_get(row, cm, 'precio', 'precio_venta', default=0) or 0)
        except: precio = 0.0
        existe = db.execute('SELECT 1 FROM lote_productos WHERE codigo=?', (codigo,)).fetchone()
        if existe:
            db.execute('UPDATE lote_productos SET descripcion=?, costo_usd=?, precio_venta=? WHERE codigo=?',
                       (desc, costo, precio, codigo))
            actualizados += 1
        else:
            db.execute(
                'INSERT INTO lote_productos (codigo, descripcion, cantidad, costo_usd, precio_venta) VALUES (?,?,1,?,?)',
                (codigo, desc, costo, precio)
            )
            insertados += 1
    db.commit()
    total = db.execute('SELECT COUNT(*) FROM lote_productos').fetchone()[0]
    return jsonify({'ok': True, 'insertados': insertados, 'actualizados': actualizados, 'total': total,
                    'mensaje': f'{insertados} nuevos, {actualizados} actualizados — {total} productos en total'})

@app.route('/api/lote/preview', methods=['POST'])
def lote_preview():
    if 'excel' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400
    archivo = request.files['excel']
    fname   = archivo.filename.lower()
    if not fname.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx, .xls)'}), 400
    try:
        if fname.endswith('.xlsx'):
            import openpyxl
            wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        else:
            import xlrd
            wb   = xlrd.open_workbook(file_contents=archivo.read())
            ws   = wb.sheet_by_index(0)
            rows = [tuple(ws.row_values(r)) for r in range(ws.nrows)]
    except Exception as e:
        return jsonify({'error': f'Error al leer el Excel: {e}'}), 400
    def to_float(v):
        try: return float(v or 0)
        except: return 0.0
    def to_int(v):
        try: return max(1, int(float(v or 1)))
        except: return 1

    if not rows:
        return jsonify({'error': 'El archivo está vacío'}), 400
    items = []
    for row in rows[1:]:
        codigo = str(_get_pos(row, 'codigo') or '').strip().upper()
        desc   = str(_get_pos(row, 'descripcion') or '').strip()
        if not codigo or not desc:
            continue
        cantidad = to_int(_get_pos(row, 'cantidad', default=1))
        costo    = to_float(_get_pos(row, 'costo', default=0))
        precio   = to_float(_get_pos(row, 'precio', default=0))
        items.append({'codigo': codigo, 'descripcion': desc,
                      'cantidad': cantidad, 'costo_usd': costo, 'precio_venta': precio})

    db = get_db()
    for item in items:
        db.execute(
            'INSERT OR REPLACE INTO lote_productos (codigo, descripcion, cantidad, costo_usd, precio_venta) VALUES (?,?,?,?,?)',
            (item['codigo'], item['descripcion'], item['cantidad'], item['costo_usd'], item['precio_venta'])
        )
    db.commit()
    return jsonify({'ok': True, 'items': items})

@app.route('/imprimir_custom', methods=['POST'])
def imprimir_custom():
    d    = request.json or {}
    cfg  = d.get('printer', {})
    modo = cfg.get('modo', 'windows')
    item = {
        'codigo':      d.get('codigo', ''),
        'descripcion': d.get('descripcion', ''),
        'cantidad':    int(d.get('cantidad', 1)),
        'referencia':  d.get('referencia', ''),
    }
    precio_venta   = d.get('precio_venta', 0)
    costo_usd      = d.get('costo_usd', 0)
    precio_etiqueta = round(float(precio_venta or 0) * 1.25, 2)
    ts  = int(datetime.utcnow().timestamp())
    zpl = generar_zpl_item(item, precio_etiqueta, costo_usd, ts)

    # Actualizar precio y costo en DB si el producto existe
    db = get_db()
    db.execute(
        'UPDATE lote_productos SET costo_usd=?, precio_venta=? WHERE codigo=?',
        (costo_usd, precio_venta, item['codigo'].strip().upper())
    )
    db.commit()

    if modo == 'red':
        ip = cfg.get('ip', '').strip()
        if not ip: return jsonify({'error': 'Debe indicar la IP'}), 400
        ok, msg = enviar_red(zpl, ip, cfg.get('puerto', 9100))
    elif modo == 'windows':
        pname = cfg.get('printer_name', '').strip()
        if not pname: return jsonify({'error': 'Debe indicar el nombre de la impresora'}), 400
        ok, msg = enviar_windows(zpl, pname)
    else:
        ok, msg = enviar_usb(zpl, cfg.get('port', 'COM1'))

    return jsonify({'ok': ok, 'mensaje': msg, 'timestamp': ts})

@app.route('/api/estanteria/preview', methods=['POST'])
def estanteria_preview():
    if 'excel' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400
    archivo = request.files['excel']
    fname   = archivo.filename.lower()
    if not fname.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Solo se aceptan archivos Excel (.xlsx, .xls)'}), 400
    try:
        if fname.endswith('.xlsx'):
            import openpyxl
            wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        else:
            import xlrd
            wb   = xlrd.open_workbook(file_contents=archivo.read())
            ws   = wb.sheet_by_index(0)
            rows = [tuple(ws.row_values(r)) for r in range(ws.nrows)]
    except Exception as e:
        return jsonify({'error': f'Error al leer el Excel: {e}'}), 400

    if not rows:
        return jsonify({'error': 'El archivo está vacío'}), 400
    cm = _col_map(rows[0])
    seen = set()
    items = []
    for row in rows[1:]:
        codigo = str(_get(row, cm, 'codigo') or '').strip().upper()
        if not codigo or codigo in seen: continue
        seen.add(codigo)
        desc = str(_get(row, cm, 'descripcion', default='') or '').strip()
        ref  = str(_get(row, cm, 'referencia',  default='') or '').strip()
        items.append({'codigo': codigo, 'descripcion': desc, 'referencia': ref, 'cantidad': 1})
    return jsonify({'ok': True, 'items': items})


@app.route('/imprimir_estanteria', methods=['POST'])
def imprimir_estanteria():
    d    = request.json or {}
    cfg  = d.get('printer', {})
    modo = cfg.get('modo', 'windows')
    codigo = str(d.get('codigo', '')).strip().upper()
    desc   = str(d.get('descripcion', '')).strip()
    ref    = str(d.get('referencia', '')).strip()
    zpl  = generar_zpl_estanteria(codigo, desc, ref, d.get('cantidad', 1))

    # Persist referencia (and descripcion) — isolated so DB errors never block printing
    if codigo:
        try:
            db = get_db()
            exists = db.execute('SELECT 1 FROM lote_productos WHERE codigo=?', (codigo,)).fetchone()
            if exists:
                db.execute('UPDATE lote_productos SET referencia=?, descripcion=? WHERE codigo=?',
                           (ref, desc or codigo, codigo))
            else:
                db.execute(
                    'INSERT INTO lote_productos (codigo, descripcion, referencia, cantidad, costo_usd, precio_venta) VALUES (?,?,?,1,0,0)',
                    (codigo, desc or codigo, ref))
            db.commit()
        except Exception as e:
            app.logger.warning('No se pudo guardar referencia para %s: %s', codigo, e)
    if modo == 'red':
        ip = cfg.get('ip', '').strip()
        if not ip: return jsonify({'error': 'Debe indicar la IP'}), 400
        ok, msg = enviar_red(zpl, ip, cfg.get('puerto', 9100))
    elif modo == 'windows':
        pname = cfg.get('printer_name', '').strip()
        if not pname: return jsonify({'error': 'Debe indicar el nombre de la impresora'}), 400
        ok, msg = enviar_windows(zpl, pname)
    else:
        ok, msg = enviar_usb(zpl, cfg.get('port', 'COM1'))
    return jsonify({'ok': ok, 'mensaje': msg})


@app.route('/api/proveedores', methods=['GET'])
def proveedores_list():
    rows = get_db().execute('SELECT codigo, nombre FROM proveedores ORDER BY nombre').fetchall()
    return jsonify({'proveedores': [{'codigo': r[0], 'nombre': r[1]} for r in rows]})

@app.route('/api/proveedores', methods=['POST'])
def proveedores_create():
    data   = request.json
    codigo = str(data.get('codigo', '')).strip().upper()
    nombre = str(data.get('nombre', '')).strip()
    if not codigo or not nombre:
        return jsonify({'ok': False, 'error': 'Código y nombre son requeridos'}), 400
    try:
        get_db().execute('INSERT INTO proveedores (codigo, nombre) VALUES (?,?)', (codigo, nombre))
        get_db().commit()
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    return jsonify({'ok': True})

@app.route('/api/proveedores/<codigo>', methods=['DELETE'])
def proveedores_delete(codigo):
    get_db().execute('DELETE FROM proveedores WHERE codigo=?', (codigo.upper(),))
    get_db().commit()
    return jsonify({'ok': True})

@app.route('/api/tasa', methods=['GET', 'POST'])
def api_tasa():
    db = get_db()
    if request.method == 'POST':
        tasa = request.json.get('tasa', 1)
        db.execute("INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('tasa_cambio', ?)", (str(tasa),))
        db.commit()
        return jsonify({'ok': True, 'tasa': tasa})
    row = db.execute("SELECT valor FROM configuracion WHERE clave='tasa_cambio'").fetchone()
    return jsonify({'tasa': float(row[0]) if row else 1.0})

def _apply_migrations():
    """Run at every startup to add new columns to existing databases."""
    db = sqlite3.connect(app.config['DATABASE'])
    try:
        db.execute("ALTER TABLE lote_productos ADD COLUMN referencia TEXT DEFAULT ''")
        db.commit()
    except Exception:
        pass  # column already exists
    db.close()

_apply_migrations()

if __name__ == '__main__':
    init_db()
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
